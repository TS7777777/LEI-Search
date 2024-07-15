from flask import Flask, request, render_template, send_file
import openpyxl
import os
from scraper import fetch_xiaohongshu_post_data, load_cookies

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files['file']
    
    if file.filename.endswith('.xlsx'):
        file_path = os.path.join('uploads', file.filename)
        file.save(file_path)
        
        # 加载Cookies
        cookies = load_cookies('cookies.json')
        
        results = process_excel(file_path, cookies)
        output_file = create_output_excel(results)
        return send_file(output_file, as_attachment=True)
    else:
        return 'Invalid file format. Please upload an Excel file.'

def process_excel(file_path, cookies):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    results = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        url = row[0]
        data = fetch_xiaohongshu_post_data(url, cookies)
        if data:
            results.append([url, data['title'], data['likes'], data['favorites'], data['comments'], data['shares']])
        else:
            results.append([url, 'Error fetching data', 'N/A', 'N/A', 'N/A', 'N/A'])
    
    return results

def create_output_excel(data):
    output_workbook = openpyxl.Workbook()
    sheet = output_workbook.active
    sheet.append(['URL', 'Title', 'Likes', 'Favorites', 'Comments', 'Shares'])

    for row in data:
        sheet.append(row)

    output_path = 'output.xlsx'
    output_workbook.save(output_path)
    return output_path

if __name__ == '__main__':
    app.run(debug=True)
